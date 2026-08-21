# -*- coding: utf-8 -*-
"""Carga los titulares reales (jueces, vocales, jefaturas) desde el reporte
oficial de "Conformación" de la CSJ Lima (ConformacionCSJLima.xlsx, hojas
SALAS / JUZGADOS / PAZ LETRADO / TRANSITORIOS / ODECMA) hacia el campo
Dependencia.titular.

Uso:
    python -m app.cargar_titulares [ruta_xlsx] [--sede "Sede Javier Alzamora Valdez"]

Sin --sede procesa todas las sedes que aparezcan en el archivo.

Formato del archivo (5 hojas, cada una con varios "bloques" de columnas en
paralelo): cada órgano ocupa una racha de líneas en una sola columna --
encabezado de categoría opcional (ej. "SALA CIVIL", se ignora), nombre
específico (ej. "2° SALA CIVIL"), "SEDE ...", "Teléfono: ... Anexo: ...", y
una o más líneas de magistrado con formato "APELLIDOS NOMBRES (T/P/S)".

Reglas (consistentes con "nunca inventar información"):
- Solo actualiza dependencias que YA existen en la base (por nombre+sede
  normalizados) -- nunca crea dependencias nuevas desde este archivo.
- Nunca sobrescribe un titular ya cargado (a mano o en una corrida previa)
  -- si el campo no está vacío, se deja tal cual.
- Lo que no logra emparejar con ninguna dependencia existente se imprime al
  final para revisión manual, no se descarta en silencio.
- (T)/(P)/(S) se interpretan como Titular/Provisional/Supernumerario --
  terminología estándar del Poder Judicial peruano; cualquier código que no
  se reconozca se deja tal cual entre paréntesis en vez de adivinar.
"""
import argparse
import re
import sys
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app import crud, models
from app.cargar_directorio_pj import formatear_nombre_sede, limpiar
from app.database import SessionLocal
from app.models import normalizar

BASE_DIR = Path(__file__).resolve().parent.parent

CONDICION = {"T": "Titular", "P": "Provisional", "S": "Supernumerario"}
_PATRON_JUEZ = re.compile(r"^(.+?)\s*\(([A-Za-zÁÉÍÓÚÑ]{1,3})\)$")

# El reporte de "Conformación" abrevia algunos nombres distinto al directorio
# oficial, aunque es el mismo órgano (mismo número, mismo título base --
# verificado comparando ambos documentos). Fallback SOLO cuando el nombre
# exacto no matchea; nunca se usa para crear dependencias nuevas.
_EQUIVALENCIAS_SUFIJO = [
    (
        "juzgado de familia subespec violencia contra la mujer e integ grup fam",
        "juzgado de familia subespecialidad en violencia contra la mujer e integrantes del grupo familiar",
    ),
    ("juzgado de trabajo", "juzgado de trabajo permanente"),
]
_EQUIVALENCIAS_EXACTAS = {
    "11 juzgado constitucional subespec temas tribut aduan indecopi y de mercado":
        "11 juzgado constitucional con subespecialidad en temas tributarios aduaneros de indecopi y de mercado",
    "jefatura de odanc": "jefatura de la odanc",
}


def _nombre_equivalente(nq: str):
    if nq in _EQUIVALENCIAS_EXACTAS:
        return _EQUIVALENCIAS_EXACTAS[nq]
    m = re.match(r"^(\d+)\s+(.*)$", nq)
    if m:
        numero, resto = m.groups()
        for patron, reemplazo in _EQUIVALENCIAS_SUFIJO:
            if resto == patron:
                return f"{numero} {reemplazo}"
    return None


def _es_sede(texto: str) -> bool:
    return texto.upper().startswith("SEDE ")


def _es_telefono(texto: str) -> bool:
    return texto.upper().startswith("TELÉFONO") or texto.upper().startswith("TELEFONO")


def _extraer_bloque(ws, columna: int, fila_inicio: int = 3) -> list[dict]:
    items = []
    for r in range(fila_inicio, ws.max_row + 1):
        v = ws.cell(row=r, column=columna).value
        if v is None:
            continue
        texto = limpiar(str(v))
        if texto:
            items.append(texto)

    organos = []
    actual = None
    i, n = 0, len(items)
    while i < n:
        texto = items[i]
        if _es_sede(texto):
            if actual is not None:
                actual["sede"] = formatear_nombre_sede(texto[len("SEDE"):].strip())
            i += 1
            continue
        if _es_telefono(texto):
            i += 1
            continue
        m = _PATRON_JUEZ.match(texto)
        if m and actual is not None:
            nombre_juez, cod = m.group(1).strip(), m.group(2).strip().upper()
            actual["jueces"].append((nombre_juez, CONDICION.get(cod, cod)))
            i += 1
            continue
        # No es sede/teléfono/magistrado: es nombre de órgano específico solo
        # si la siguiente línea no vacía es "SEDE ..."; si no, es un
        # encabezado de categoría (ej. "SALA CIVIL") y se ignora.
        if i + 1 < n and _es_sede(items[i + 1]):
            if actual is not None:
                organos.append(actual)
            actual = {"nombre": texto, "sede": None, "jueces": []}
        i += 1
    if actual is not None:
        organos.append(actual)
    return organos


def extraer_organos(origen) -> list[dict]:
    """origen: ruta (str/Path) o archivo en memoria (BytesIO) -- openpyxl
    acepta ambos indistintamente."""
    wb = load_workbook(origen, data_only=True)
    organos = []
    for nombre_hoja in wb.sheetnames:
        ws = wb[nombre_hoja]
        columnas = sorted({
            celda.column
            for fila in ws.iter_rows(min_row=3)
            for celda in fila
            if celda.value not in (None, "")
        })
        for columna in columnas:
            organos.extend(_extraer_bloque(ws, columna))
    return organos


def _formatear_titular(jueces: list[tuple[str, str]]) -> str:
    partes = []
    for nombre, condicion in jueces:
        nombre_legible = " ".join(p.capitalize() for p in nombre.split())
        partes.append(f"{nombre_legible} ({condicion})")
    return "; ".join(partes)


def aplicar(db: Session, organos: list[dict], sede_filtro: str | None) -> dict:
    """Aplica los titulares detectados y devuelve un resumen -- lo usa tanto
    el comando de línea (que lo imprime) como el endpoint del panel admin
    (que lo devuelve como JSON)."""
    sedes_por_nombre = {normalizar(s.nombre): s for s in db.query(models.Sede).all()}
    sede_norm_filtro = normalizar(sede_filtro) if sede_filtro else None

    actualizadas = 0
    ya_tenian = 0
    sin_emparejar: list[str] = []

    for organo in organos:
        if not organo["sede"] or not organo["jueces"]:
            continue
        if sede_norm_filtro and normalizar(organo["sede"]) != sede_norm_filtro:
            continue

        sede = sedes_por_nombre.get(normalizar(organo["sede"]))
        if not sede:
            sin_emparejar.append(f'{organo["nombre"]} -- la sede "{organo["sede"]}" no existe en el catálogo')
            continue

        nq = normalizar(organo["nombre"])
        dep = (
            db.query(models.Dependencia)
            .filter(models.Dependencia.sede_id == sede.id, models.Dependencia.nombre_normalizado == nq)
            .first()
        )
        if not dep:
            alt = _nombre_equivalente(nq)
            if alt:
                dep = (
                    db.query(models.Dependencia)
                    .filter(models.Dependencia.sede_id == sede.id, models.Dependencia.nombre_normalizado == alt)
                    .first()
                )
        if not dep:
            sin_emparejar.append(f'{organo["nombre"]} ({organo["sede"]}) -- no está en el catálogo de esa sede')
            continue

        if dep.titular:
            ya_tenian += 1
            continue

        dep.titular = _formatear_titular(organo["jueces"])
        actualizadas += 1

    db.commit()
    return {"actualizadas": actualizadas, "ya_tenian": ya_tenian, "sin_emparejar": sin_emparejar}


def _xlsx_por_defecto() -> Path:
    candidato = BASE_DIR / "ConformacionCSJLima.xlsx"
    if not candidato.exists():
        raise SystemExit(f"No se encontró {candidato}")
    return candidato


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("ruta", nargs="?", default=None)
    parser.add_argument("--sede", default=None, help='Ej.: "Sede Javier Alzamora Valdez". Si se omite, procesa todas.')
    args = parser.parse_args()

    ruta = Path(args.ruta) if args.ruta else _xlsx_por_defecto()
    print(f"Leyendo: {ruta}")
    organos = extraer_organos(str(ruta))
    print(f"Órganos detectados en el archivo: {len(organos)}")

    db = SessionLocal()
    try:
        resumen = aplicar(db, organos, args.sede)
        crud.auditoria.registrar(
            db, "sistema", "dependencia", None, "IMPORTAR_TITULARES",
            f"Carga masiva desde {ruta.name} (vía terminal): {resumen['actualizadas']} actualizadas, "
            f"{resumen['ya_tenian']} ya tenían titular, {len(resumen['sin_emparejar'])} sin emparejar"
            + (f" (sede: {args.sede})" if args.sede else " (todas las sedes)") + ".",
        )
    finally:
        db.close()

    print(f"Dependencias actualizadas con titular: {resumen['actualizadas']}")
    print(f"Ya tenían titular (no se tocaron): {resumen['ya_tenian']}")
    print(f"Sin emparejar en el catálogo ({len(resumen['sin_emparejar'])}):")
    for linea in resumen["sin_emparejar"]:
        print(f"  - {linea}")
