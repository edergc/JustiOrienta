"""Importa la plantilla Excel de levantamiento del catálogo hacia la base de datos.

Requiere que el esquema ya exista (alembic upgrade head).

Uso:
    python -m app.import_excel "JusticiaOrienta_04_Plantilla_Catalogo_Piloto.xlsx"

Empareja dependencias por 'Nombre oficial': si ya existe, la actualiza; si no,
la crea. Las sedes y edificios mencionados por nombre se crean automáticamente
si todavía no existen (para no obligar a levantarlos aparte antes de poder
cargar el catálogo). Las filas sin nombre se ignoran.
"""
import sys
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app import crud, models
from app.database import SessionLocal

COLUMNAS = [
    "id_hoja", "tipo", "categoria", "nombre", "alias",
    "sede", "edificio", "piso", "oficina", "horario",
    "servicios", "requisitos", "telefono", "correo",
    "rampa", "ascensor", "banio_accesible", "ruta_accesible",
    "estado", "responsable_validar",
]

SI_NO = {"sí": True, "si": True, "no": False, "no aplica": False, "": False}


def _bool(v) -> bool:
    return SI_NO.get(str(v or "").strip().lower(), False)


def _texto(v) -> str:
    return str(v).strip() if v is not None else ""


def _obtener_o_crear_sede(db: Session, nombre: str) -> models.Sede:
    sede = db.query(models.Sede).filter(models.Sede.nombre == nombre).first()
    if sede:
        return sede
    sede = models.Sede(nombre=nombre, estado="activo")
    db.add(sede)
    db.flush()
    return sede


def _obtener_o_crear_edificio(db: Session, sede_id: int, nombre: str) -> models.Edificio:
    edificio = (
        db.query(models.Edificio)
        .filter(models.Edificio.sede_id == sede_id, models.Edificio.nombre == nombre)
        .first()
    )
    if edificio:
        return edificio
    edificio = models.Edificio(sede_id=sede_id, nombre=nombre, estado="activo")
    db.add(edificio)
    db.flush()
    return edificio


def importar(ruta_excel: str) -> None:
    wb = load_workbook(ruta_excel, data_only=True)
    ws = wb["Catálogo"]

    db = SessionLocal()
    creadas, actualizadas, omitidas = 0, 0, 0

    try:
        for fila in ws.iter_rows(min_row=2, values_only=True):
            valores = dict(zip(COLUMNAS, fila))
            nombre = _texto(valores.get("nombre"))
            if not nombre:
                omitidas += 1
                continue

            tipo_map = {"jurisdiccional": "jurisdiccional", "administrativa": "administrativa", "servicio": "servicio"}
            tipo = tipo_map.get(_texto(valores.get("tipo")).lower(), "servicio")
            estado_map = {"activo": "activo", "en revisión": "revision", "en revision": "revision", "inactivo": "inactivo"}
            estado = estado_map.get(_texto(valores.get("estado")).lower(), "revision")

            nombre_sede = _texto(valores.get("sede"))
            sede_id = None
            edificio_id = None
            if nombre_sede:
                sede = _obtener_o_crear_sede(db, nombre_sede)
                sede_id = sede.id
                nombre_edificio = _texto(valores.get("edificio"))
                if nombre_edificio:
                    edificio_id = _obtener_o_crear_edificio(db, sede_id, nombre_edificio).id

            if sede_id is None:
                print(f"  omitida '{nombre}': no tiene sede indicada (columna 'Sede' vacía)")
                omitidas += 1
                continue

            data = {
                "tipo": tipo,
                "categoria": _texto(valores.get("categoria")) or None,
                "nombre": nombre,
                "sede_id": sede_id,
                "edificio_id": edificio_id,
                "piso": _texto(valores.get("piso")) or None,
                "oficina": _texto(valores.get("oficina")) or None,
                "horario": _texto(valores.get("horario")) or None,
                "servicios": _texto(valores.get("servicios")) or None,
                "requisitos": _texto(valores.get("requisitos")) or None,
                "telefono": _texto(valores.get("telefono")) or None,
                "correo": _texto(valores.get("correo")) or None,
                "rampa": _bool(valores.get("rampa")),
                "ascensor": _bool(valores.get("ascensor")),
                "banio_accesible": _bool(valores.get("banio_accesible")),
                "ruta_accesible": _bool(valores.get("ruta_accesible")),
                "estado": estado,
                "area": _texto(valores.get("categoria")) or _texto(valores.get("tipo")),
                "responsable_validar": _texto(valores.get("responsable_validar")) or None,
            }
            alias_csv = _texto(valores.get("alias"))

            existente = db.query(models.Dependencia).filter(models.Dependencia.nombre == nombre).first()
            if existente:
                crud.dependencias.actualizar(db, existente, data, alias_csv)
                actualizadas += 1
            else:
                crud.dependencias.crear(db, data, alias_csv)
                creadas += 1

        db.commit()
        print(f"Importación completa: {creadas} creadas, {actualizadas} actualizadas, {omitidas} filas omitidas.")
    finally:
        db.close()


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Uso: python -m app.import_excel <ruta_al_excel.xlsx>")
        sys.exit(1)
    ruta = Path(sys.argv[1])
    if not ruta.exists():
        print(f"No se encontró el archivo: {ruta}")
        sys.exit(1)
    importar(str(ruta))
