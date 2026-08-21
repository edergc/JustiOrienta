# -*- coding: utf-8 -*-
"""Carga el directorio real de la CSJ Lima desde el Excel oficial (formato
distinto al PDF de app/cargar_directorio_pj.py, pero mismo contenido de
fondo: sede, dirección, central telefónica, dependencia, piso, anexo).

Uso:
    python -m app.cargar_directorio_excel [ruta_al_xlsx]

Si no se indica ruta, usa DirectorioCSJLI.xlsx en la raíz del proyecto.

Reutiliza las reglas y funciones ya probadas de cargar_directorio_pj.py
(misma función cargar(), mismo formateo de nombres de sede, misma lista de
"Mesa de Partes" ya desambiguadas) -- lo único que cambia es cómo se leen
las filas de origen, porque el Excel trae "PISO"/"ANEXO" en la misma fila
que "DEPENDENCIAS JURISDICCIONALES"/"DEPENDENCIAS ADMINISTRATIVAS" en vez de
en una fila aparte como el PDF.
"""
import sys
from pathlib import Path

from openpyxl import load_workbook

from app import crud
from app.cargar_directorio_pj import YA_DESAMBIGUADAS, cargar, formatear_nombre_sede, limpiar
from app.database import SessionLocal

BASE_DIR = Path(__file__).resolve().parent.parent


def extraer_registros(ruta_xlsx: str):
    """Devuelve (sedes: {nombre: {direccion, central}}, filas: [dict]) --
    mismo formato que produce app/cargar_directorio_pj.py, para poder
    reutilizar su función cargar() sin cambios."""
    wb = load_workbook(ruta_xlsx, data_only=True)
    ws = wb.active

    sedes: dict[str, dict] = {}
    filas: list[dict] = []
    sede_actual = None
    tipo_actual = None

    for row in ws.iter_rows(min_row=1, values_only=True):
        c0 = limpiar(row[0]) if len(row) > 0 and row[0] is not None else ""
        c1 = row[1] if len(row) > 1 else None
        c2 = row[2] if len(row) > 2 else None

        if not c0:
            continue
        if c0.upper().startswith("NOTA"):
            # Pie de página ("NOTA: Actualizado al ..."), no es una fila de datos.
            break

        if c0.upper().startswith("SEDE "):
            sede_actual = formatear_nombre_sede(c0[len("SEDE"):].strip())
            tipo_actual = None
            continue

        if c0.upper().startswith("DIRECCI") and sede_actual:
            direccion = c0.split(":", 1)[-1].strip()
            central = limpiar(str(c1)) if c1 else ""
            if sede_actual not in sedes:
                sedes[sede_actual] = {"direccion": direccion, "central": central}
            continue

        if c0.upper() in ("DEPENDENCIAS JURISDICCIONALES", "DEPENDENCIAS ADMINISTRATIVAS"):
            tipo_actual = "administrativa" if "ADMINISTRATIVA" in c0.upper() else "jurisdiccional"
            continue

        if tipo_actual is None or sede_actual is None:
            continue

        piso = limpiar(str(c1)) if c1 is not None else ""
        anexo = limpiar(str(c2)) if c2 is not None else ""
        if c0 == "Mesa de Partes" and (sede_actual, piso, anexo) in YA_DESAMBIGUADAS:
            continue

        filas.append({"sede": sede_actual, "tipo": tipo_actual, "nombre": c0, "piso": piso, "anexo": anexo})

    return sedes, filas


def _xlsx_por_defecto() -> Path:
    candidato = BASE_DIR / "DirectorioCSJLI.xlsx"
    if not candidato.exists():
        raise SystemExit(f"No se encontró {candidato}")
    return candidato


if __name__ == "__main__":
    ruta = Path(sys.argv[1]) if len(sys.argv) > 1 else _xlsx_por_defecto()
    print(f"Leyendo: {ruta}")
    sedes, filas = extraer_registros(str(ruta))
    print(f"Sedes detectadas: {len(sedes)}")
    print(f"Filas de dependencias detectadas: {len(filas)}")

    db = SessionLocal()
    try:
        creadas = cargar(db, sedes, filas)
        print(f"Dependencias nuevas creadas: {creadas}")
        crud.auditoria.registrar(
            db, "sistema", "dependencia", None, "IMPORTAR_DIRECTORIO",
            f"Carga masiva desde {ruta.name}: {creadas} dependencias nuevas "
            f"(de {len(filas)} filas leídas en {len(sedes)} sedes). Solo se crea lo que no existía; "
            f"nada se sobrescribió.",
        )
    finally:
        db.close()
