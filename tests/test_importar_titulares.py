import io

from fastapi.testclient import TestClient
from openpyxl import Workbook

from app.main import app

client = TestClient(app)

_MEDIA_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _login(dni: str, password: str) -> str:
    r = client.post("/api/v1/auth/login", data={"username": dni, "password": password})
    assert r.status_code == 200, r.text
    return r.json()["access_token"]


def _auth(token: str) -> dict:
    return {"Authorization": f"Bearer {token}"}


def _xlsx_conformacion(sede_nombre: str) -> bytes:
    """Un solo bloque en la hoja SALAS, con el mismo formato que produce el
    reporte real de Conformación: categoría (se ignora), órgano específico,
    sede, teléfono/anexo, y un magistrado."""
    resto = sede_nombre[5:] if sede_nombre.lower().startswith("sede ") else sede_nombre
    wb = Workbook()
    ws = wb.active
    ws.title = "SALAS"
    ws["B3"] = "SALA CIVIL"
    ws["B4"] = "1° Juzgado Civil"
    ws["B5"] = f"SEDE {resto.upper()}"
    ws["B6"] = "Teléfono: 410 1818 Anexo: 13000"
    ws["B7"] = "PEREZ GOMEZ MARIA JOSE (T)"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def test_admin_importa_titulares_y_actualiza_dependencia_existente(db, sede, admin_usuario):
    from app import crud

    dep = crud.dependencias.crear(
        db,
        dict(tipo="jurisdiccional", nombre="1° Juzgado Civil", sede_id=sede.id, area="X", estado="activo"),
        "",
    )

    token = _login("10000001", "clave123")
    r = client.post(
        "/api/v1/admin/dependencias/importar-titulares",
        headers=_auth(token),
        files={"archivo": ("conformacion.xlsx", _xlsx_conformacion(sede.nombre), _MEDIA_XLSX)},
        data={"sede": sede.nombre},
    )
    assert r.status_code == 200, r.text
    data = r.json()
    assert data["actualizadas"] == 1
    assert data["ya_tenian"] == 0
    assert data["sin_emparejar"] == []

    db.refresh(dep)
    assert dep.titular == "Perez Gomez Maria Jose (Titular)"


def test_no_sobrescribe_titular_ya_cargado(db, sede, admin_usuario):
    from app import crud

    dep = crud.dependencias.crear(
        db,
        dict(
            tipo="jurisdiccional", nombre="1° Juzgado Civil", sede_id=sede.id, area="X", estado="activo",
            titular="Alguien Que Ya Estaba (Titular)",
        ),
        "",
    )

    token = _login("10000001", "clave123")
    r = client.post(
        "/api/v1/admin/dependencias/importar-titulares",
        headers=_auth(token),
        files={"archivo": ("conformacion.xlsx", _xlsx_conformacion(sede.nombre), _MEDIA_XLSX)},
        data={"sede": sede.nombre},
    )
    assert r.status_code == 200, r.text
    data = r.json()
    assert data["actualizadas"] == 0
    assert data["ya_tenian"] == 1

    db.refresh(dep)
    assert dep.titular == "Alguien Que Ya Estaba (Titular)"


def test_organo_sin_dependencia_existente_queda_sin_emparejar(db, sede, admin_usuario):
    token = _login("10000001", "clave123")
    r = client.post(
        "/api/v1/admin/dependencias/importar-titulares",
        headers=_auth(token),
        files={"archivo": ("conformacion.xlsx", _xlsx_conformacion(sede.nombre), _MEDIA_XLSX)},
        data={"sede": sede.nombre},
    )
    assert r.status_code == 200
    data = r.json()
    assert data["actualizadas"] == 0
    assert len(data["sin_emparejar"]) == 1


def test_gestor_no_puede_importar_titulares(sede, gestor_usuario):
    token = _login("10000002", "clave123")
    r = client.post(
        "/api/v1/admin/dependencias/importar-titulares",
        headers=_auth(token),
        files={"archivo": ("conformacion.xlsx", _xlsx_conformacion(sede.nombre), _MEDIA_XLSX)},
    )
    assert r.status_code == 403
