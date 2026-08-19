import io

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app import models
from app.main import app

client = TestClient(app)


def _token_admin():
    return client.post(
        "/api/v1/auth/login", data={"username": "10000001", "password": "clave123"}
    ).json()["access_token"]


def test_top_sin_resultado_solo_incluye_busquedas_no_encontradas(db, sede, admin_usuario):
    from app import crud

    crud.dependencias.crear(
        db,
        dict(tipo="administrativa", nombre="Recursos Humanos", sede_id=sede.id, area="Recursos Humanos", estado="activo"),
        "",
    )

    client.get("/api/v1/buscar", params={"q": "recursos humanos"})  # sí se encuentra
    client.get("/api/v1/buscar", params={"q": "certificado de antecedentes penales"})  # no existe
    client.get("/api/v1/buscar", params={"q": "certificado de antecedentes penales"})  # de nuevo

    token = _token_admin()
    resumen = client.get(
        "/api/v1/admin/metricas/resumen", headers={"Authorization": f"Bearer {token}"}
    ).json()

    consultas_sin_resultado = {t["consulta"]: t["veces"] for t in resumen["top_consultas_sin_resultado"]}
    assert consultas_sin_resultado.get("certificado de antecedentes penales") == 2
    assert "recursos humanos" not in consultas_sin_resultado


def test_pendientes_por_area_cuenta_solo_lo_que_esta_en_revision(db, sede, admin_usuario):
    db.add(models.Dependencia(
        tipo="administrativa", nombre="Pendiente 1", sede_id=sede.id, area="Recursos Humanos", estado="revision",
    ))
    db.add(models.Dependencia(
        tipo="administrativa", nombre="Pendiente 2", sede_id=sede.id, area="Recursos Humanos", estado="revision",
    ))
    db.add(models.Dependencia(
        tipo="administrativa", nombre="Ya Publicada", sede_id=sede.id, area="Recursos Humanos", estado="activo",
    ))
    db.commit()

    token = _token_admin()
    resumen = client.get(
        "/api/v1/admin/metricas/resumen", headers={"Authorization": f"Bearer {token}"}
    ).json()

    assert resumen["pendientes_total"] == 2
    assert resumen["pendientes_antiguedad_promedio_dias"] is not None
    fila = next(p for p in resumen["pendientes_por_area"] if p["area"] == "Recursos Humanos")
    assert fila["cantidad"] == 2


def test_sin_pendientes_el_indicador_no_falla(admin_usuario):
    token = _token_admin()
    resumen = client.get(
        "/api/v1/admin/metricas/resumen", headers={"Authorization": f"Bearer {token}"}
    ).json()
    assert resumen["pendientes_total"] == 0
    assert resumen["pendientes_antiguedad_promedio_dias"] is None
    assert resumen["pendientes_por_area"] == []


def test_reporte_xlsx_incluye_las_hojas_nuevas(db, sede, admin_usuario):
    db.add(models.Dependencia(
        tipo="administrativa", nombre="Pendiente", sede_id=sede.id, area="Recursos Humanos", estado="revision",
    ))
    db.commit()
    client.get("/api/v1/buscar", params={"q": "algo que no existe en el catalogo"})

    token = _token_admin()
    r = client.get("/api/v1/admin/metricas/reporte.xlsx", headers={"Authorization": f"Bearer {token}"})
    assert r.status_code == 200

    wb = load_workbook(io.BytesIO(r.content))
    assert "Búsquedas sin resultado" in wb.sheetnames
    assert "Pendientes por área" in wb.sheetnames
    assert "Completitud por área" in wb.sheetnames


def test_completitud_por_area_solo_cuenta_lo_publicado(db, sede, admin_usuario):
    # Completa: horario, teléfono y rampa -- 3/3 campos.
    db.add(models.Dependencia(
        tipo="administrativa", nombre="Completa", sede_id=sede.id, area="Recursos Humanos", estado="activo",
        horario="8 a 16", telefono="111", rampa=True,
    ))
    # Sin nada de eso -- 0/3 campos.
    db.add(models.Dependencia(
        tipo="administrativa", nombre="Incompleta", sede_id=sede.id, area="Recursos Humanos", estado="activo",
    ))
    # En revisión: no debe contar para nada, ni siquiera aunque esté completa.
    db.add(models.Dependencia(
        tipo="administrativa", nombre="Aun No Publicada", sede_id=sede.id, area="Recursos Humanos",
        estado="revision", horario="8 a 16", telefono="111", rampa=True,
    ))
    db.commit()

    token = _token_admin()
    resumen = client.get(
        "/api/v1/admin/metricas/resumen", headers={"Authorization": f"Bearer {token}"}
    ).json()

    fila = next(c for c in resumen["completitud_por_area"] if c["area"] == "Recursos Humanos")
    assert fila["activas"] == 2  # no cuenta la que está en revisión
    # (3 campos completos + 0 campos) / (2 dependencias * 3 campos) = 50%
    assert fila["porcentaje_completo"] == 50.0
    assert resumen["porcentaje_completitud_global"] == 50.0


def test_completitud_sin_dependencias_activas_no_falla(admin_usuario):
    token = _token_admin()
    resumen = client.get(
        "/api/v1/admin/metricas/resumen", headers={"Authorization": f"Bearer {token}"}
    ).json()
    assert resumen["completitud_por_area"] == []
    assert resumen["porcentaje_completitud_global"] is None
