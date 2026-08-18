import io

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.import_excel import COLUMNAS
from app.main import app

client = TestClient(app)

_POS = {nombre: i for i, nombre in enumerate(COLUMNAS)}


def _login(dni: str, password: str) -> str:
    r = client.post("/api/v1/auth/login", data={"username": dni, "password": password})
    assert r.status_code == 200, r.text
    return r.json()["access_token"]


def _auth(token: str) -> dict:
    return {"Authorization": f"Bearer {token}"}


def test_exportar_catalogo_usa_las_mismas_posiciones_que_import_excel(db, sede, admin_usuario):
    from app import crud

    crud.dependencias.crear(
        db,
        dict(
            tipo="jurisdiccional", categoria="Civil", nombre="1er Juzgado Civil", sede_id=sede.id,
            piso="3", oficina="301", horario="8 a 16", telefono="111", area="Juzgado civil",
            estado="activo", rampa=True,
        ),
        "alias uno, alias dos",
    )

    atoken = _login("10000001", "clave123")
    r = client.get("/api/v1/admin/dependencias/exportar.xlsx", headers=_auth(atoken))
    assert r.status_code == 200
    assert r.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    wb = load_workbook(io.BytesIO(r.content))
    assert "Catálogo" in wb.sheetnames
    ws = wb["Catálogo"]
    fila = [c.value for c in ws[2]]  # fila 1 es encabezado

    # app.import_excel.importar() empareja columnas por POSICIÓN (zip con
    # COLUMNAS), no por el texto del encabezado -- así que lo que importa acá
    # es que cada dato caiga en el índice correcto, no que el encabezado diga
    # literalmente "tipo".
    assert fila[_POS["tipo"]] == "jurisdiccional"
    assert fila[_POS["categoria"]] == "Civil"
    assert fila[_POS["nombre"]] == "1er Juzgado Civil"
    assert fila[_POS["sede"]] == sede.nombre
    assert fila[_POS["piso"]] == "3"
    assert fila[_POS["oficina"]] == "301"
    assert fila[_POS["horario"]] == "8 a 16"
    assert fila[_POS["telefono"]] == "111"
    assert fila[_POS["rampa"]] == "Sí"
    assert fila[_POS["estado"]] == "activo"
    assert "alias uno" in fila[_POS["alias"]]
    assert "alias dos" in fila[_POS["alias"]]


def test_exportar_catalogo_incluye_revision_e_inactivo_no_solo_activo(db, sede, admin_usuario):
    from app import crud

    crud.dependencias.crear(
        db, dict(tipo="administrativa", nombre="En Revision", sede_id=sede.id, area="X", estado="revision"), "",
    )
    crud.dependencias.crear(
        db, dict(tipo="administrativa", nombre="Ya Inactiva", sede_id=sede.id, area="X", estado="inactivo"), "",
    )

    atoken = _login("10000001", "clave123")
    r = client.get("/api/v1/admin/dependencias/exportar.xlsx", headers=_auth(atoken))
    wb = load_workbook(io.BytesIO(r.content))
    ws = wb["Catálogo"]
    nombres = [row[_POS["nombre"]] for row in ws.iter_rows(min_row=2, values_only=True)]
    assert "En Revision" in nombres
    assert "Ya Inactiva" in nombres


def test_exportar_catalogo_distingue_sin_confirmar_de_no(db, sede, admin_usuario):
    """Si rampa/ascensor quedan en NULL en la base (columna nullable -- p. ej.
    datos corregidos a mano, o cargados antes de que este campo existiera),
    exportar eso como "No" afirmaría que se confirmó la falta de accesibilidad,
    cuando en realidad nadie la revisó. Se fuerza el NULL con un UPDATE en vez
    de pasarlo al crear la dependencia: una asignación None a nivel de
    instancia ORM cae en el default=False de la columna (comportamiento
    estándar de SQLAlchemy), así que no sirve para simular este estado."""
    from sqlalchemy import update

    from app import crud, models

    dep = crud.dependencias.crear(
        db, dict(tipo="administrativa", nombre="Recien Cargada", sede_id=sede.id, area="X", estado="revision"), "",
    )
    db.execute(update(models.Dependencia).where(models.Dependencia.id == dep.id).values(rampa=None, ascensor=None))
    db.commit()

    atoken = _login("10000001", "clave123")
    r = client.get("/api/v1/admin/dependencias/exportar.xlsx", headers=_auth(atoken))
    wb = load_workbook(io.BytesIO(r.content))
    ws = wb["Catálogo"]
    fila = next(row for row in ws.iter_rows(min_row=2, values_only=True) if row[_POS["nombre"]] == "Recien Cargada")
    assert fila[_POS["rampa"]] == "Sin confirmar"
    assert fila[_POS["ascensor"]] == "Sin confirmar"


def test_gestor_exporta_solo_su_propia_area(db, sede, gestor_usuario):
    from app import crud

    crud.dependencias.crear(
        db, dict(tipo="administrativa", nombre="De Recursos Humanos", sede_id=sede.id,
                 area="Recursos Humanos", estado="activo"), "",
    )
    crud.dependencias.crear(
        db, dict(tipo="administrativa", nombre="De Otra Area", sede_id=sede.id,
                 area="Otra Área", estado="activo"), "",
    )

    gtoken = _login("10000002", "clave123")
    r = client.get("/api/v1/admin/dependencias/exportar.xlsx", headers=_auth(gtoken))
    assert r.status_code == 200
    wb = load_workbook(io.BytesIO(r.content))
    ws = wb["Catálogo"]
    nombres = [row[_POS["nombre"]] for row in ws.iter_rows(min_row=2, values_only=True)]
    assert "De Recursos Humanos" in nombres
    assert "De Otra Area" not in nombres


def test_consulta_no_puede_exportar_el_catalogo(consulta_usuario):
    token = _login("10000004", "clave123")
    r = client.get("/api/v1/admin/dependencias/exportar.xlsx", headers=_auth(token))
    assert r.status_code == 403
